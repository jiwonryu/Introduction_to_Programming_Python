#1
import openpyxl

filename = "sample.xlsx"

wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
for i in range(1,max_row+1):
    for j in range(1,max_col+1):
        currentCell_obj = sheet_obj.cell(i,j)
        print(currentCell_obj.value, end=' ')
    print()

#2
newCell_obj = sheet_obj['A5']
newCell_obj.value = 'Brian'

newCell_obj = sheet_obj['B5']
newCell_obj.value = '5'

newCell_obj = sheet_obj['C5']
newCell_obj.value = '3489203'
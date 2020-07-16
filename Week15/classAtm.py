from classBankAccount import BankAccount #classBankAccount 모듈 부터 BankAccount class 받아옴
import time #시간 관련 모듈

class Atm (BankAccount): #class 상속
    def __init__(self, name='None', balance=0): #인사 안내말 없애기
        self.name = name
        self.balance = balance
    def balancecheck(self): # 잔액 확인해주는 함수
        print('현재 잔액은 ', self.balance, '입니다')
    def receipt(self): # 명세표 출력 관련 함수
        print("명세표를 출력하시겠습니까?(y/n) ", end=' ')
        while True:
            y_or_n = input()
            if y_or_n == "y": # y선택시 명세표 출력
                print("*"*40)
                print("{0:^40}".format("명세표"))
                print("거래 시간: ", time.asctime(time.localtime(time.time()))) # 현재 시각 보여주는 함수
                print("이름:", username)
                if menu == 1: #입금 선택했을 때는 입금액 표시
                    print("입금액:", DepositAmount)
                elif menu == 2: #출금 선택했을 때는 출금액 표시
                    print("출금액:", WithdrawAmount)
                print("남은 잔액:", self.balance)
                print("거래해주셔서 감사합니다. - by Python Bank")
                print("*"*40)
                break
            elif y_or_n == "n": # n 선택시 그냥 넘어감
                break
            else: # 다른 문자 입력시 안내 메시지 출력
                print("잘못된 입력입니다. 명세표를 출력하시려면 y, 출력하지 않으시려면 n을 입력해주세요.", end=' ')

import openpyxl # 엑셀 사용 관련 모듈
filename = "UserList.xlsx"
wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active

max_row = sheet_obj.max_row

UserList = {'name':('pw','balance')} # 엑셀에 있는 UesrList 불러와 딕셔너리로 정리하기
for i in range(2,max_row+1): # 엑셀의 두번째 row부터 마지막 row까지
    usernameCell = sheet_obj.cell(i,1).value #column 1에 해당하는 라인은 사용자 이름
    userpwCell = sheet_obj.cell(i,2).value #column 2에 해당하는 라인은 사용자 비밀번호
    userbalanceCell = sheet_obj.cell(i,3).value #column 3에 해당하는 라인은 사용자 잔액
    UserList[usernameCell] = (userpwCell,userbalanceCell) #UserList 딕셔너리에 'Key : Value = 이름: (비밀번호, 잔액)'으로  추가

print("♥ Python Bank에 오신 것을 환영합니다. ")

while True:
    username = input("사용자의 이름을 입력해주세요: ") # 사용자 이름 받기
    if username in UserList: #리스트에 입력한 이름이 있으면
        print("%s 님 환영합니다." %username)
        break
    else: #리스트에 입력한 이름이 없으면
        print("%s 님은 등록되지 않았습니다. 추가하시겠습니까?(yes/no)" %username, end = ' ')
        while True:
            yes_or_no = input()
            if yes_or_no == "yes": # yes를 입력할 경우
                newpassword = int(input("%s 님의 비밀번호를 입력하세요: " %username))
                newbalance = int(input("%s 님의 초기 잔액을 입력하세요: " %username))
                sheet_obj.cell(max_row+1,1).value = username #엑셀에 이름 추가
                sheet_obj.cell(max_row+1,2).value = newpassword #엑셀에 비밀번호 추가
                sheet_obj.cell(max_row+1,3).value = newbalance #엑셀에 잔액 추가
                wb_obj.save(filename) #엑셀파일 저장
                UserList[username] = (newpassword, newbalance) #딕셔너리에 추가
                print("등록이 완료되었습니다!")
                break
            elif yes_or_no == "no": # no를 입력할 경우
                print("다시", end = ' ')
                break
            else:
                print("잘못된 입력입니다. 사용자를 등록하시려면 yes, 등록하시지 않으려면 no를 입력해주세요.", end=' ')

userpassword = UserList.get(username)[0] #해당 사용자의 비밀번호
userbalance = UserList.get(username)[1] #해당 사용자의 잔액

UseAtm = Atm(username, userbalance) # 해당 사용자의 이름, 잔액으로 인스턴스 생성

print("%s 님의 비밀번호를 입력하세요: " %username, end=' ') #비밀번호 입력받는 메시지
while True:
    password = int(input()) #비밀번호 입력받음
    if password == userpassword:  # 엑셀파일에 등록된 비밀번호(딕셔너리의 value)와 매치되면        print("사용자 정보가 확인되었습니다.")
        break
    else: #엑셀파일에 등록된 비밀번호가 아니면
        print("비밀번호가 틀렸습니다. 비밀번호를 다시 입력해주세요.", end= ' ') #다시 비밀번호 입력하도록 함

while True: # 메뉴 반복되게끔 함
    print("=" * 40)
    print("원하시는 메뉴를 선택하세요.")
    print("1. Deposit")
    print("2. Withdraw")
    print("3. Check Balance")
    print("4. Quit")
    menu = int(input()) #원하는 메뉴를 입력 받음

    if menu == 1: #입금 선택시
        DepositAmount = int(input("입금하실 금액을 입력하세요: "))
        UseAtm.deposit(DepositAmount) #입금한 금액을 입력받아 UseAtm의 deposit 함수 이용
        UseAtm.receipt() #명세표 출력 관련 함수

    elif menu == 2: #출금
        WithdrawAmount = int(input("출금하실 금액을 입력하세요: "))
        UseAtm.withdraw(WithdrawAmount) #출금한 금액을 입력받아 UseAtm의 withdraw 함수 이용
        UseAtm.receipt() #명세표 출력 관련 함수

    elif menu == 3: #잔액조회
        UseAtm.balancecheck() #잔액 조회 관련 함수

    elif menu == 4: #나가기
        print("이용해주셔서 감사합니다. 좋은 하루 되세요. - Python Bank")
        break

    else:
        print("없는 메뉴입니다. 다시 선택해 주세요.")
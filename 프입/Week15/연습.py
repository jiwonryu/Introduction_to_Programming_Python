from classBankAccount import BankAccount
import time

class Atm (BankAccount):
    def balancecheck():
        print('현재 잔액은 ', self.balance, '입니다')
    def receipt():
        #명세표 출력 관련 함수
        y_or_n = str(input ("명세표를 출력하시겠습니까?(y/n) "))
        if y_or_n == "y":
            print("*"*30)
            print("{0:^30}".format("명세표"))
            print("거래 시간: ", time.asctime(time.localtime(time.time()))) #시간 재는 코드
            print("이름:", username)
            print("입금액:", dpstamount)
            print("남은 잔액:", self.balance)
            print("거래해주셔서 감사합니다. - by Python Bank")
            print("*"*30)
    #def continueBank:
        #a = int(input("거래를 계속하시길 원하면 1, 종료하시러면 2를 눌러주세요: "))
        #if a == 1:
        #    menu = True
        #    return menu
        #elif a == 2:
        #    menu = False
        #    return menu


import openpyxl
filename = "UserList.xlsx"
wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active
#엑셀열어서 리스트 확인 웅앵
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

UserList = {'name':('pw','balance')}

for i in range(2,max_row):
    usernameCell = sheet_obj.cell(i,1).value
    userpwCell = sheet_obj.cell(i,2).value
    userbalanceCell = sheet_obj.cell(i,3).value
    UserList[usernameCell] = (userpwCell,userbalanceCell)

username = input("Enter the username: ")
if username in UserList:
    print(username, "님 환영합니다.")
else: #리스트에 등록안돼있으면
    yes_or_no = input("%s 님은 등록되지 않았습니다. 추가하시겠습니까?(yes/no)" %username)
    if yes_or_no == "yes":
        newpassword = int(input("%s 님의 비밀번호를 입력하세요: " %username))
        newbalance = int(input("%s 님의 초기 잔액을 입력하세요: " %username))
        sheet_obj.cell(1,max_row+1).value = username
        sheet_obj.cell(2,max_row+1).value = newpassword
        sheet_obj.cell(3,max_row+1).value = newbalance
        print("등록이 완료되었습니다!")
    elif yes_or_no == "no":
        pass

print(UserList)
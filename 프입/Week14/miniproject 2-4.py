#2-4
from mod_scoregrade import PythonLecture
#앞에서 만든 모듈의 PythonLecture Class 불러옴

import openpyxl #엑셀모듈

filename = "score_table.xlsx" #score_table.xlsx"파일로 부터 데이터 받아옴

wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active

newCell_obj = sheet_obj['F1']
newCell_obj.value = 'Final Score' # F1칸에 'Final Score'추가
newCell_obj = sheet_obj['G1']
newCell_obj.value = 'Final Grade' # G1칸에 'Final Grade'추가

max_row = sheet_obj.max_row
for i in range(2, max_row+1): # row2부터 마지막 row까지 아래 내용 반복
    fst = (sheet_obj.cell(i,2)).value # 두번째 칸에 해당하는 값은 1차 점수
    scd = (sheet_obj.cell(i,3)).value # 세번째 칸에 해당하는 값은 2차 점수
    thrd = (sheet_obj.cell(i,4)).value # 네번째 칸에 해당하는 값은 3차 점수
    absnc = (sheet_obj.cell(i,5)).value # 다섯번째 칸에 해당하는 값은 결석 횟수
    std = PythonLecture(fst,scd,thrd,absnc)
    # 위 네개의 값을 PythonLecture class에 대입한 인스턴스 생성

    score_obj = sheet_obj.cell(i,6) # 여섯번째 칸에
    score_obj.value = std.score() # 최종 점수 입력
    grade_obj = sheet_obj.cell(i,7) # 일곱번째 칸에
    garde_obj.value = std.grade() # 학점 입력

wb_obj.save('score_table_final.xlsx')
#'score_table_final.xlsx'라는 새로운 엑셀파일 생성


